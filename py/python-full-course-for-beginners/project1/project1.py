# openpyxl is a Python library used to read from and write to Microsoft Excel files
import openpyxl as xl  # "as xl" is not required, but enables "xl. ..." instead of "openpyxl. ..."
from openpyxl.chart import BarChart, Reference  # Importing two classes from chart module from openpyxyl library


def process_workbook(filename):  # Reusable function for ANY file
    wb = xl.load_workbook(filename)
    sheet = wb["Sheet1"]

    # Both lines return the same cell
    # cell = sheet["A1"]
    # cell = sheet.cell(1, 1)
    # print(cell.value)
    # print(sheet.max_row) <- Returns number of rows containing data

    for row in range(2, sheet.max_row + 1):  # range() doesn't naturally include 2nd input, so adding 1 ensures it does
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)  # Selecting cells in column 4 in rows 2 to 4

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, "A6")
    # It's good practice to save file under new name to not accidentally overwrite original file if code breaks
    wb.save(filename)  # Can't have Excel file open while running code

process_workbook("transactions2.xlsx")